"""
1.1 메타데이터 API 라우터

엔드포인트 (기존):
  GET    /dashboard              - 오늘 통계
  GET    /contents               - 콘텐츠 목록 (페이징·필터)
  POST   /contents               - 콘텐츠 수동 등록
  GET    /contents/{id}          - 콘텐츠 상세
  POST   /contents/{id}/process  - AI 처리 트리거
  GET    /queue                  - 검수 큐 (70~89점)
  POST   /queue/{id}/action      - 검수 액션 (승인/수정/반려)
  POST   /generate               - 실시간 메타 AI 생성
  GET    /emails                 - CP 이메일 수신 이력

신규 엔드포인트 (파이프라인):
  GET    /staging                    - 검토 대기풀 목록
  POST   /staging/bulk-approve       - 벌크 승인
  POST   /staging/bulk-reject        - 벌크 반려
  POST   /contents/{id}/enrich       - 에이전틱 검색 수동 트리거
  GET    /contents/{id}/hierarchy    - 시리즈 계층 트리
  GET    /pipeline/status            - 파이프라인 현황
  POST   /upload/batch               - CSV/엑셀 배치 업로드
  GET    /upload/batch/{job_id}      - 배치 작업 상태 조회

신규 엔드포인트 (3분류 메타):
  GET    /text                       - 글자메타 목록 (시리즈 계층 포함)
  GET    /text/{id}                  - 특정 콘텐츠 글자메타
  PUT    /text/{id}                  - 글자메타 수정
  POST   /text/bulk-complete         - 글자메타 일괄 완료
  GET    /image                      - 이미지메타 목록
  GET    /image/{id}                 - 특정 콘텐츠 이미지 목록
  GET    /video                      - 영상메타 목록
  GET    /video/{id}                 - 특정 콘텐츠 영상메타
  PUT    /video/{id}                 - 영상메타 수정
  POST   /video/bulk-complete        - 영상메타 일괄 완료
  GET    /service-readiness          - 서비스 준비 현황 통계
"""

import io
import csv
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session

from shared.database import get_db
from api.programming.metadata import service
from api.programming.metadata.models import ContentStatus, ContentType, ContentBatchJob
from api.programming.metadata.schemas import (
    ContentCreate, ContentUpdate, ContentOut, ContentDetail, PaginatedContents,
    MetadataReviewAction, AIGenerateRequest, AIGenerateResponse,
    DashboardStats, CpEmailLogOut,
    PaginatedStagingItems, StagingItem, BulkActionRequest, PipelineStatus,
    BatchJobOut,
    TextMetaOut, TextMetaUpdate, TextMetaBulkCompleteRequest, TextMetaSuggestion,
    ImageMetaOut, ImageBulkCompleteRequest, ImageMetaSuggestions,
    VideoMetaOut, VideoMetaUpdate, VideoBulkCompleteRequest,
    ServiceReadinessStats,
    PaginatedTmdbItems,
    TmdbCacheStats, TmdbSyncLogItem, PaginatedSyncLog, TmdbCacheRecentItem,
    ExternalSourceStats, PaginatedExternalItems,
    BulkActionConsolidatedRequest, BulkActionResponse, JobStatusOut,
    UndoActionRequest, UndoActionOut,
    PromoteAIResultOut, ApplyExternalFieldsRequest,
    ContentChangelogOut, LockFieldsRequest,
    EnrichPreviewRequest, EnrichPreviewOut, BatchPreviewOut, SourceSearchOut, CreateFromSourcesRequest, CreateFromSourcesOut,
    PosterCandidateOut, PosterRecommendResponse, PosterSelectRequest,
    RecommendationsOut,
    PaginatedAiReviewQueue,
)
from api.programming.metadata.models import CpEmailLog
from api.programming.metadata import poster_recommend

router = APIRouter()


@router.get("/dashboard", response_model=DashboardStats)
def get_dashboard(db: Session = Depends(get_db)):
    return service.get_dashboard_stats(db)


@router.get("/contents", response_model=PaginatedContents)
def list_contents(
    status: ContentStatus | None = Query(None),
    cp_name: str | None = Query(None),
    title: str | None = Query(None),
    content_type: ContentType | None = Query(None),
    production_year: int | None = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    items, total = service.list_contents(
        db,
        status=status,
        cp_name=cp_name,
        title=title,
        content_type=content_type,
        production_year=production_year,
        page=page,
        size=size,
    )
    result = []
    for c in items:
        out = ContentOut.model_validate(c)
        if c.metadata_record:
            out.quality_score = c.metadata_record.quality_score
        out.poster_url = service._primary_poster_url(c)
        result.append(out)
    return PaginatedContents(items=result, total=total, page=page, size=size)


@router.post("/contents", response_model=ContentOut, status_code=201)
def create_content(data: ContentCreate, db: Session = Depends(get_db)):
    content = service.create_content(db, data)
    return ContentOut.model_validate(content)


@router.get("/contents/{content_id}", response_model=ContentDetail)
def get_content(content_id: int, db: Session = Depends(get_db)):
    content = service.get_content(db, content_id)
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    out = ContentDetail.model_validate(content)
    out.poster_url = service._primary_poster_url(content)
    if content.metadata_record:
        out.quality_score = content.metadata_record.quality_score
    return out


@router.put("/contents/{content_id}", response_model=ContentOut)
def update_content(content_id: int, data: ContentUpdate, db: Session = Depends(get_db)):
    """수동 수정 — 입력 필드를 manual source로 저장 후 resolve_metadata 재실행"""
    try:
        content = service.update_content(db, content_id, data)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return ContentOut.model_validate(content)


@router.post("/contents/{content_id}/process")
def trigger_processing(content_id: int, db: Session = Depends(get_db)):
    content = service.get_content(db, content_id)
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    task_id = service.trigger_ai_processing(content_id)
    return {"task_id": task_id, "message": "AI 처리 큐에 등록되었습니다"}


@router.get("/queue", response_model=PaginatedContents)
def get_review_queue(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    items, total = service.get_review_queue(db, page=page, size=size)
    result = []
    for c in items:
        out = ContentOut.model_validate(c)
        if c.metadata_record:
            out.quality_score = c.metadata_record.quality_score
        result.append(out)
    return PaginatedContents(items=result, total=total, page=page, size=size)


@router.post("/queue/{content_id}/action", response_model=ContentOut)
def review_action(
    content_id: int,
    action: MetadataReviewAction,
    db: Session = Depends(get_db),
):
    try:
        content = service.apply_review_action(db, content_id, action)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return ContentOut.model_validate(content)


@router.post("/generate", response_model=AIGenerateResponse)
async def generate_metadata(
    req: AIGenerateRequest,
    db: Session = Depends(get_db),
):
    """실시간 메타 AI 생성 — Ollama llama3.2:3b 호출"""
    from api.programming.metadata.ai_engine import generate_metadata_ollama
    result = await generate_metadata_ollama(req, db)
    return result


@router.get("/emails", response_model=list[CpEmailLogOut])
def list_email_logs(
    processed: bool | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    q = db.query(CpEmailLog)
    if processed is not None:
        q = q.filter(CpEmailLog.processed == processed)
    return q.order_by(CpEmailLog.received_at.desc()).limit(limit).all()


# ── Staging 대기풀 엔드포인트 ─────────────────────────────

@router.get("/staging", response_model=PaginatedStagingItems)
def get_staging_queue(
    content_type: Optional[str] = Query(None, description="movie | series | season | episode"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """운영자 검토 대기풀 목록 (staging 상태, 최상위 콘텐츠 + 계층 포함)"""
    items, total = service.get_staging_queue(db, content_type=content_type, page=page, size=size)
    return PaginatedStagingItems(items=items, total=total, page=page, size=size)


@router.post("/staging/bulk-approve")
def bulk_approve(req: BulkActionRequest, db: Session = Depends(get_db)):
    """선택 항목 일괄 승인 (staging → approved)"""
    return service.bulk_approve_staging(db, req)


@router.post("/staging/bulk-reject")
def bulk_reject(req: BulkActionRequest, db: Session = Depends(get_db)):
    """선택 항목 일괄 반려 (staging → rejected)"""
    return service.bulk_reject_staging(db, req)


# ── 에이전틱 검색 수동 트리거 ─────────────────────────────

@router.post("/contents/{content_id}/enrich")
def trigger_enrichment(content_id: int, db: Session = Depends(get_db)):
    """에이전틱 멀티소스 검색 수동 트리거 (TMDB + KOBIS)"""
    content = service.get_content(db, content_id)
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    task_id = service.trigger_enrichment(content_id)
    return {"task_id": task_id, "message": "에이전틱 검색 큐에 등록되었습니다"}


@router.post("/contents/{content_id}/recommend-posters", response_model=PosterRecommendResponse)
def recommend_posters(content_id: int, db: Session = Depends(get_db)):
    """TMDB /images API 에서 포스터 후보를 수집해 ContentImage 에 멱등 upsert."""
    content = service.get_content(db, content_id)
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    images, added = poster_recommend.recommend_posters_for_content(db, content_id)
    return PosterRecommendResponse(
        content_id=content_id,
        candidates=[PosterCandidateOut.model_validate(img) for img in images],
        added=added,
    )


@router.get("/contents/{content_id}/poster-candidates", response_model=list[PosterCandidateOut])
def get_poster_candidates(content_id: int, db: Session = Depends(get_db)):
    """콘텐츠의 poster 타입 ContentImage 전체 반환 (is_primary DESC, id ASC)."""
    content = service.get_content(db, content_id)
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    from api.programming.metadata.models import ContentImage, ImageType
    images = (
        db.query(ContentImage)
        .filter(
            ContentImage.content_id == content_id,
            ContentImage.image_type == ImageType.poster,
        )
        .order_by(ContentImage.is_primary.desc(), ContentImage.id.asc())
        .all()
    )
    return [PosterCandidateOut.model_validate(img) for img in images]


@router.post("/contents/{content_id}/poster/select", response_model=list[PosterCandidateOut])
def select_poster(content_id: int, req: PosterSelectRequest, db: Session = Depends(get_db)):
    """image_id 의 poster 를 primary 로 지정, 나머지는 primary 해제."""
    content = service.get_content(db, content_id)
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    try:
        poster_recommend.select_primary_poster(db, content_id, req.image_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    from api.programming.metadata.models import ContentImage, ImageType
    images = (
        db.query(ContentImage)
        .filter(
            ContentImage.content_id == content_id,
            ContentImage.image_type == ImageType.poster,
        )
        .order_by(ContentImage.is_primary.desc(), ContentImage.id.asc())
        .all()
    )

    from shared.config import settings
    from workers.tasks.metadata import send_dam_webhook
    if settings.DAM_POSTER_INGEST_URL:
        selected = next((img for img in images if img.is_primary), None)
        if selected:
            from datetime import datetime, timezone
            send_dam_webhook.delay(
                event_type="poster_primary_set",
                content_id=content_id,
                title=content.title,
                content_type=content.content_type.value if hasattr(content.content_type, "value") else str(content.content_type),
                occurred_at=datetime.now(timezone.utc).isoformat(),
                poster_url=selected.url,
                poster_source=selected.source or "tmdb",
                image_id=selected.id,
            )

    return [PosterCandidateOut.model_validate(img) for img in images]


@router.get("/contents/{content_id}/hierarchy", response_model=StagingItem)
def get_content_hierarchy(content_id: int, db: Session = Depends(get_db)):
    """시리즈 계층 트리 반환 (시리즈 > 시즌 > 에피소드)"""
    item = service.get_content_hierarchy(db, content_id)
    if not item:
        raise HTTPException(status_code=404, detail="Content not found")
    return item


@router.get("/contents/{content_id}/recommendations", response_model=RecommendationsOut)
def get_recommendations(content_id: int, db: Session = Depends(get_db)):
    """빈 메타 필드 감지 + 외부소스·AI 추천값 반환 (메타 보강 제안 패널용)"""
    return service.get_content_recommendations(db, content_id)


# ── 파이프라인 현황 ────────────────────────────────────────

@router.get("/pipeline/status", response_model=PipelineStatus)
def get_pipeline_status(db: Session = Depends(get_db)):
    """파이프라인 현황: 상태별 콘텐츠 수, 평균 품질, 실패 항목"""
    return service.get_pipeline_status(db)


# ── 배치 업로드 ───────────────────────────────────────────

@router.post("/upload/batch", response_model=BatchJobOut, status_code=201)
async def batch_upload(
    file: UploadFile = File(..., description="CSV 또는 Excel 파일"),
    cp_name: Optional[str] = Form(None),
    created_by: Optional[str] = Form(None),
    dry_run: Optional[bool] = Query(False, description="건강성 검사만 수행 (job 생성 스킵)"),
    db: Session = Depends(get_db),
):
    """CSV/엑셀 배치 업로드 → 파싱 후 Content(waiting) 생성 + AI 처리 큐 등록"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="파일명이 없습니다")

    allowed_extensions = (".csv", ".xlsx", ".xls")
    if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
        raise HTTPException(status_code=400, detail="CSV 또는 Excel 파일만 허용됩니다")

    content_bytes = await file.read()
    file_size = len(content_bytes)
    if file_size > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(status_code=400, detail="파일 크기가 10MB를 초과합니다")

    # 배치 작업 생성
    job = service.create_batch_job(
        db, file_name=file.filename, cp_name=cp_name,
        created_by=created_by, file_size=file_size,
    )

    # CSV 파싱 (Excel은 추후 openpyxl 연동)
    rows = []
    try:
        def _extract_row(get_fn) -> dict:
            """공통 필드 추출 — CSV/Excel 모두 사용"""
            return {
                "title": get_fn(["title", "제목"]),
                "production_year": _safe_int(get_fn(["production_year", "제작연도"])),
                "content_type": _normalize_content_type(get_fn(["content_type", "타입"]) or "movie"),
                "cp_name": get_fn(["cp_name", "CP사"]) or cp_name,
                "synopsis": get_fn(["synopsis", "시놉시스"]),
                "cast": get_fn(["cast", "출연진"]),
                "directors": get_fn(["directors", "감독"]),
                "genres": get_fn(["genres", "장르"]),
                "country": get_fn(["country", "제작국가"]),
                "runtime": _safe_int(get_fn(["runtime", "런타임"])),
                "rating_age": get_fn(["rating_age", "시청등급"]),
                "poster_url": get_fn(["poster_url", "포스터URL"]) or None,
            }

        if file.filename.lower().endswith(".csv"):
            text = content_bytes.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                def _csv_get(col_names, _row=row):
                    for n in col_names:
                        v = _row.get(n)
                        if v:
                            return v.strip()
                    return ""
                rows.append(_extract_row(_csv_get))
        else:
            # Excel 지원 — openpyxl 없으면 에러 메시지 반환
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
                header_map = {h: i for i, h in enumerate(headers)}
                for row_cells in ws.iter_rows(min_row=2, values_only=True):
                    def _excel_get(col_names, _cells=row_cells, _hm=header_map):
                        for n in col_names:
                            idx = _hm.get(n)
                            if idx is not None and _cells[idx]:
                                return str(_cells[idx]).strip()
                        return ""
                    rows.append(_extract_row(_excel_get))
            except ImportError:
                raise HTTPException(
                    status_code=422,
                    detail="Excel 파일 처리에 openpyxl이 필요합니다. CSV를 사용하거나 서버에 openpyxl을 설치하세요."
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"파일 파싱 실패: {exc}")

    # 비어있는 행 제거
    rows = [r for r in rows if r.get("title")]

    result = service.process_batch_rows(db, job, rows)
    db.refresh(job)
    return BatchJobOut.model_validate(job)


@router.get("/upload/batch/{job_id}", response_model=BatchJobOut)
def get_batch_job(job_id: int, db: Session = Depends(get_db)):
    """배치 작업 상태 조회"""
    job = db.query(ContentBatchJob).filter(ContentBatchJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="배치 작업을 찾을 수 없습니다")
    return BatchJobOut.model_validate(job)


# ── 글자메타 엔드포인트 ───────────────────────────────────────────────────────

@router.get("/text", summary="글자메타 목록 (시리즈 계층 포함)")
def list_text_meta(
    completed: Optional[bool] = Query(None, description="완료 여부 필터 (true/false)"),
    content_type: Optional[str] = Query(None, description="movie | series"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    return service.get_text_meta_list(db, completed=completed, content_type_filter=content_type, page=page, size=size)


@router.post("/text/bulk-complete", summary="글자메타 일괄 완료 처리")
def bulk_complete_text_meta(req: TextMetaBulkCompleteRequest, db: Session = Depends(get_db)):
    return service.bulk_complete_text_meta(db, req.content_ids)


@router.get("/text/{content_id}", response_model=TextMetaOut, summary="특정 콘텐츠 글자메타")
def get_text_meta(content_id: int, db: Session = Depends(get_db)):
    result = service.get_text_meta(db, content_id)
    if not result:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


@router.put("/text/{content_id}", response_model=TextMetaOut, summary="글자메타 수정")
def update_text_meta(content_id: int, data: TextMetaUpdate, db: Session = Depends(get_db)):
    result = service.update_text_meta(db, content_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


# ── 글자메타 AI 제안 ──────────────────────────────────────────────────────────

@router.get("/text/{content_id}/suggest", response_model=TextMetaSuggestion, summary="글자메타 AI 제안")
def suggest_text_meta(content_id: int, db: Session = Depends(get_db)):
    """TMDB/KOBIS 외부 소스 기반 글자메타 제안. 없으면 기존 AI 결과 반환."""
    result = service.suggest_text_meta(db, content_id)
    if not result:
        raise HTTPException(status_code=404, detail="제안 데이터가 없습니다. TMDB/KOBIS 동기화 또는 AI 처리를 먼저 실행하세요.")
    return result


# ── 이미지메타 엔드포인트 ─────────────────────────────────────────────────────

@router.get("/image", summary="이미지메타 목록")
def list_image_meta(
    completed: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    return service.get_image_meta_list(db, completed=completed, page=page, size=size)


@router.post("/image/bulk-complete", summary="이미지메타 일괄 완료 처리")
def bulk_complete_image_meta(req: ImageBulkCompleteRequest, db: Session = Depends(get_db)):
    return service.bulk_complete_image_meta(db, req.content_ids)


@router.get("/image/{content_id}/suggest", response_model=ImageMetaSuggestions, summary="TMDB 이미지 제안")
def suggest_image_meta(content_id: int, db: Session = Depends(get_db)):
    """TMDB 외부 소스에서 누락 이미지 타입별 URL 제안"""
    result = service.suggest_image_meta(db, content_id)
    if result is None:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


@router.post("/image/{content_id}/upload", response_model=ImageMetaOut, status_code=201, summary="이미지 URL 등록")
def upload_image(
    content_id: int,
    image_type: str = Form(..., description="poster|thumbnail|stillcut|banner|logo"),
    url: str = Form(..., description="이미지 URL"),
    width: Optional[int] = Form(None),
    height: Optional[int] = Form(None),
    source: str = Form("manual", description="cp|tmdb|manual"),
    db: Session = Depends(get_db),
):
    """이미지 URL을 ContentImage로 등록. 5종 완료 시 image_meta_completed 자동 갱신."""
    try:
        result = service.add_content_image(db, content_id, image_type, url, width, height, source)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    if not result:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


@router.get("/image/{content_id}", response_model=ImageMetaOut, summary="특정 콘텐츠 이미지 목록")
def get_image_meta(content_id: int, db: Session = Depends(get_db)):
    result = service.get_image_meta(db, content_id)
    if not result:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


# ── 영상메타 엔드포인트 ───────────────────────────────────────────────────────

@router.get("/video", summary="영상메타 목록")
def list_video_meta(
    completed: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    return service.get_video_meta_list(db, completed=completed, page=page, size=size)


@router.post("/video/bulk-complete", summary="영상메타 일괄 완료 처리")
def bulk_complete_video_meta(req: VideoBulkCompleteRequest, db: Session = Depends(get_db)):
    return service.bulk_complete_video_meta(db, req.content_ids)


@router.get("/video/{content_id}", response_model=VideoMetaOut, summary="특정 콘텐츠 영상메타")
def get_video_meta(content_id: int, db: Session = Depends(get_db)):
    result = service.get_video_meta(db, content_id)
    if not result:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


@router.put("/video/{content_id}", response_model=VideoMetaOut, summary="영상메타 수정")
def update_video_meta(content_id: int, data: VideoMetaUpdate, db: Session = Depends(get_db)):
    result = service.update_video_meta(db, content_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="콘텐츠를 찾을 수 없습니다")
    return result


# ── 서비스 준비 현황 ──────────────────────────────────────────────────────────

@router.get("/service-readiness", response_model=ServiceReadinessStats, summary="서비스 준비 현황 통계")
def get_service_readiness(db: Session = Depends(get_db)):
    return service.get_service_readiness(db)


# ── TMDB 동기화 결과 ──────────────────────────────────────────────────────────

@router.get("/tmdb", response_model=PaginatedTmdbItems, summary="TMDB 매핑 콘텐츠 목록")
def list_tmdb_synced(
    content_type: Optional[str] = Query(None, description="movie | series"),
    search: Optional[str] = Query(None, description="제목 검색"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    items, total = service.list_tmdb_synced(db, content_type=content_type, search=search, page=page, size=size)
    return PaginatedTmdbItems(items=items, total=total, page=page, size=size)


def _safe_int(val) -> Optional[int]:
    try:
        return int(val) if val else None
    except (ValueError, TypeError):
        return None


def _normalize_content_type(val: str) -> str:
    mapping = {
        "영화": "movie", "movie": "movie",
        "시리즈": "series", "series": "series", "드라마": "series",
        "시즌": "season", "season": "season",
        "에피소드": "episode", "episode": "episode",
    }
    return mapping.get(val.strip(), "movie")


# ── TMDB 캐시 모니터링 ────────────────────────────────────────────────────────

@router.get("/tmdb-cache/stats", response_model=TmdbCacheStats, summary="TMDB 캐시 통계")
def get_tmdb_cache_stats(db: Session = Depends(get_db)):
    return service.get_tmdb_cache_stats(db)


@router.get("/tmdb-cache/sync-log", response_model=PaginatedSyncLog, summary="TMDB 동기화 로그")
def list_tmdb_sync_log(
    source: Optional[str] = Query(None, description="discover_movie | changes_movie | backfill_movie_year 등"),
    status: Optional[str] = Query(None, description="running | completed | failed"),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    items, total = service.list_tmdb_sync_log(db, source=source, status=status, page=page, size=size)
    return PaginatedSyncLog(items=items, total=total, page=page, size=size)


@router.get("/tmdb-cache/recent", response_model=list[TmdbCacheRecentItem], summary="최근 fetch 캐시 항목")
def list_tmdb_cache_recent(
    kind: str = Query("both", description="movie | tv | both"),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    return service.list_tmdb_cache_recent(db, kind=kind, limit=limit)


# ── KOBIS 모니터링 ────────────────────────────────────────────────────────────

@router.get("/kobis/stats", response_model=ExternalSourceStats, summary="KOBIS 동기화 통계")
def get_kobis_stats(db: Session = Depends(get_db)):
    return service.get_external_source_stats(db, "kobis")


@router.get("/kobis/sync-log", response_model=PaginatedSyncLog, summary="KOBIS 동기화 로그")
def list_kobis_sync_log(
    status: Optional[str] = Query(None, description="running | completed | failed"),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    items, total = service.list_external_source_sync_log(db, "kobis", status=status, page=page, size=size)
    return PaginatedSyncLog(items=items, total=total, page=page, size=size)


@router.get("/kobis/search", response_model=PaginatedExternalItems, summary="KOBIS 캐시 검색")
def search_kobis(
    title: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    items, total = service.search_external_sources(db, "kobis", title=title, year=None, page=page, size=size)
    return PaginatedExternalItems(items=items, total=total, page=page, size=size)


# ── KMDB 모니터링 ────────────────────────────────────────────────────────────

@router.get("/kmdb/stats", response_model=ExternalSourceStats, summary="KMDB 동기화 통계")
def get_kmdb_stats(db: Session = Depends(get_db)):
    return service.get_external_source_stats(db, "kmdb")


@router.get("/kmdb/sync-log", response_model=PaginatedSyncLog, summary="KMDB 동기화 로그 (항상 빈 결과)")
def list_kmdb_sync_log(
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    items, total = service.list_external_source_sync_log(db, "kmdb", status=status, page=page, size=size)
    return PaginatedSyncLog(items=items, total=total, page=page, size=size)


@router.get("/kmdb/search", response_model=PaginatedExternalItems, summary="KMDB 캐시 검색")
def search_kmdb(
    title: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    items, total = service.search_external_sources(db, "kmdb", title=title, year=None, page=page, size=size)
    return PaginatedExternalItems(items=items, total=total, page=page, size=size)


# ── dev-api-consolidation: Bulk Actions ──────────────────────

@router.post("/bulk/reprocess", response_model=BulkActionResponse, status_code=202, summary="Bulk AI 재처리")
async def api_bulk_reprocess(
    req: BulkActionConsolidatedRequest,
    db: Session = Depends(get_db),
):
    import os
    sync_mode = os.environ.get("BULK_SYNC_MODE", "false").lower() == "true"
    return await service.bulk_reprocess(db, req.ids, req.reason, sync_mode)


@router.post("/bulk/enrich", response_model=BulkActionResponse, status_code=202, summary="Bulk 외부 재매칭")
async def api_bulk_enrich(
    req: BulkActionConsolidatedRequest,
    db: Session = Depends(get_db),
):
    import os
    sync_mode = os.environ.get("BULK_SYNC_MODE", "false").lower() == "true"
    return await service.bulk_enrich(db, req.ids, req.reason, sync_mode)


@router.post("/bulk/process", response_model=BulkActionResponse, status_code=202, summary="Bulk 즉시 처리")
async def api_bulk_process(
    req: BulkActionConsolidatedRequest,
    db: Session = Depends(get_db),
):
    import os
    sync_mode = os.environ.get("BULK_SYNC_MODE", "false").lower() == "true"
    return await service.bulk_process(db, req.ids, req.reason, sync_mode)


@router.post("/bulk/recall", response_model=BulkActionResponse, status_code=202, summary="Bulk 회수")
async def api_bulk_recall(
    req: BulkActionConsolidatedRequest,
    db: Session = Depends(get_db),
):
    import os
    sync_mode = os.environ.get("BULK_SYNC_MODE", "false").lower() == "true"
    return await service.bulk_recall(db, req.ids, req.reason, sync_mode)


@router.delete("/bulk", response_model=BulkActionResponse, status_code=202, summary="Bulk soft delete")
async def api_bulk_delete(
    req: BulkActionConsolidatedRequest,
    db: Session = Depends(get_db),
):
    import os
    sync_mode = os.environ.get("BULK_SYNC_MODE", "false").lower() == "true"
    return await service.bulk_delete(db, req.ids, req.reason, sync_mode)


@router.get("/contents/jobs/{job_id}", response_model=JobStatusOut, summary="Job 상태 조회")
async def api_get_job_status(job_id: int, db: Session = Depends(get_db)):
    return await service.get_job_status(db, job_id)


@router.post("/bulk/undo", response_model=UndoActionOut, summary="Bulk 액션 되돌리기")
async def api_bulk_undo(req: UndoActionRequest, db: Session = Depends(get_db)):
    return await service.bulk_undo(db, req.action_id)


@router.post("/contents/jobs/{job_id}/retry-failed", response_model=BulkActionResponse, status_code=202, summary="실패 항목 재실행")
async def api_retry_failed_in_job(job_id: int, db: Session = Depends(get_db)):
    return await service.retry_failed_in_job(db, job_id)


@router.post("/contents/{content_id}/ai-results/{result_id}/promote", response_model=PromoteAIResultOut, summary="AI 결과 채택")
async def api_promote_ai_result(content_id: int, result_id: int, db: Session = Depends(get_db)):
    return await service.promote_ai_result(db, content_id, result_id)


@router.post("/contents/{content_id}/process", response_model=JobStatusOut, status_code=202, summary="부분 필드 재처리")
async def api_partial_reprocess(
    content_id: int,
    fields: str = Query(""),
    db: Session = Depends(get_db),
):
    field_list = [f.strip() for f in fields.split(",") if f.strip()]
    return await service.partial_reprocess(db, content_id, field_list)


@router.post("/contents/{content_id}/external/{source_id}/apply-fields", summary="외부 필드 적용")
async def api_apply_external_fields(
    content_id: int,
    source_id: int,
    req: ApplyExternalFieldsRequest,
    db: Session = Depends(get_db),
):
    return await service.apply_external_fields(db, content_id, source_id, req.fields)


@router.get("/contents/{content_id}/changelog", response_model=ContentChangelogOut, summary="변경 이력 조회")
async def api_get_changelog(
    content_id: int,
    db: Session = Depends(get_db),
):
    return await service.get_changelog(db, content_id)


@router.get("/ai-review-queue", response_model=PaginatedAiReviewQueue, summary="AI 검수 큐 — 메타+포스터+Dam 통합")
def api_get_ai_review_queue(
    status: Optional[str] = None,
    input_type: Optional[str] = None,
    metadata_status: Optional[str] = None,
    poster_status: Optional[str] = None,
    risk_level: Optional[str] = None,
    include_dam: bool = False,
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    return service.build_ai_review_queue(
        db,
        status=status,
        input_type=input_type,
        metadata_status=metadata_status,
        poster_status=poster_status,
        risk_level=risk_level,
        include_dam=include_dam,
        page=page,
        size=size,
    )


@router.post("/contents/{content_id}/lock", summary="필드 잠금")
async def api_lock_fields(
    content_id: int,
    req: LockFieldsRequest,
    db: Session = Depends(get_db),
):
    return await service.lock_fields(db, content_id, req.fields, req.reason)


@router.post("/contents/{content_id}/preview-clip", status_code=202, summary="Preview clip 생성 요청")
async def api_request_preview_clip(
    content_id: int,
    db: Session = Depends(get_db),
):
    return await service.request_preview_clip(db, content_id)


@router.post("/contents/{content_id}/enrich", response_model=EnrichPreviewOut, summary="Enrich 미리보기")
async def api_enrich_preview(
    content_id: int,
    preview: bool = Query(False),
    db: Session = Depends(get_db),
):
    if preview:
        return await service.enrich_preview(db, content_id)
    return {"error": "preview=true 필수"}


@router.get("/sources/search", response_model=SourceSearchOut, summary="소스 통합 검색")
async def api_sources_search(
    q: str = Query(...),
    sources: str = Query("tmdb,kobis"),
    db: Session = Depends(get_db),
):
    source_list = [s.strip() for s in sources.split(",")]
    return await service.sources_search(db, q, source_list)


@router.post("/contents/from_sources", response_model=CreateFromSourcesOut, summary="소스에서 콘텐츠 생성")
async def api_create_from_sources(
    req: CreateFromSourcesRequest,
    db: Session = Depends(get_db),
):
    return await service.create_from_sources(db, req.source_id, req.selected_fields, req.cp_name)


@router.post("/contents/{content_id}/enrich-credits", summary="외부 소스에서 credits 보강")
async def api_enrich_credits(
    content_id: int,
    db: Session = Depends(get_db),
):
    return await service.enrich_external_credits(content_id, db)
